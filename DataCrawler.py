from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
import re
row_to_erase=34

def split_data_with_negative_signs(data):
    if isinstance(data, str):
        return re.findall(r'-?\d+', data)
    else:
        return []

driver = webdriver.Chrome()
url = 'https://wdc.kugi.kyoto-u.ac.jp/dst_realtime/202405/dst2405.for.request'

driver.get(url)

wait = WebDriverWait(driver, 10)  
wait.until(EC.presence_of_element_located((By.TAG_NAME, 'pre')))  

html_content = driver.page_source

soup = BeautifulSoup(html_content, 'html.parser')

data_tag = soup.find('pre')

if data_tag:
    data_text = data_tag.text.strip()
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Data'
    
    data_lines = data_text.splitlines()
    for idx, line in enumerate(data_lines, start=2):
        ws.cell(row=idx, column=1).value = line.strip()
    
    wb.save('raw_data.xlsx')
    driver.quit()  

    input_excel_file = 'raw_data.xlsx'
    output_excel_file = 'cooked_data.xlsx'
    
    df = pd.read_excel(input_excel_file)
    
    df['Numeric_Data'] = df['Data'].apply(split_data_with_negative_signs)
    
    numeric_data = pd.DataFrame(df['Numeric_Data'].tolist(), columns=[f'Value_{i}' for i in range(1, df['Numeric_Data'].apply(len).max() + 1)])
    
    df = pd.concat([df, numeric_data], axis=1)
    
    df.to_excel(output_excel_file, index=False)
output_excel_file = 'cooked_data.xlsx'
df = pd.read_excel(output_excel_file)

df.drop(columns=['Value_1', 'Value_3'], inplace=True)

df.rename(columns={'Value_2': 'Data_instance', 'Value_4': 'Value_1'}, inplace=True)

new_column_order = ['Data', 'Data_instance', 'Value_1', 'Value_5', 'Value_6', 'Value_7', 'Value_8', 'Value_9', 'Value_10', 'Value_11', 'Value_12', 'Value_13', 'Value_14', 'Value_15', 'Value_16', 'Value_17', 'Value_18', 'Value_19', 'Value_20', 'Value_21', 'Value_22', 'Value_23', 'Value_24', 'Value_25']

df = df[new_column_order]
start_column = 'Value_5'
start_index = df.columns.get_loc(start_column)

new_column_names = {df.columns[i]: f'Value_{i - start_index + 2}' for i in range(start_index, len(df.columns))}

df.rename(columns=new_column_names, inplace=True)
df.drop(index=32, inplace=True)

df.to_excel(output_excel_file, index=False)